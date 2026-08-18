from __future__ import annotations

import logging
from pathlib import Path

from app.document.parsers.base import BaseParser, DocumentSection

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parse Excel (.xlsx) files using openpyxl."""

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]

    def parse(self, file_path: str) -> list[DocumentSection]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed")
            return [DocumentSection(
                title=Path(file_path).stem,
                content="[Excel 解析需要安装 openpyxl: pip install openpyxl]",
                metadata={"file": file_path, "parser": "fallback"},
            )]

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sections: list[DocumentSection] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(cells):  # skip empty rows
                    rows.append(cells)

            if not rows:
                continue

            # Header row as title
            header = rows[0] if rows else []
            title = f"工作表: {sheet_name}"

            # Convert to natural language for embedding
            content_lines = []
            if len(rows) > 1:
                for row in rows[1:min(51, len(rows))]:  # limit to 50 data rows
                    pairs = []
                    for i, cell in enumerate(row):
                        if cell and i < len(header) and header[i]:
                            pairs.append(f"{header[i]}: {cell}")
                    if pairs:
                        content_lines.append("；".join(pairs))

            content = "\n".join(content_lines) if content_lines else " | ".join(header)

            sections.append(DocumentSection(
                title=title,
                content=content,
                section_type="table",
                page_number=None,
                metadata={
                    "sheet_name": sheet_name,
                    "row_count": len(rows),
                    "col_count": len(header),
                    "headers": [str(h) for h in header],
                    "raw_rows": rows[:51],  # store first 50 data rows
                },
            ))

        wb.close()
        return sections if sections else [DocumentSection(content="[空工作簿]")]
